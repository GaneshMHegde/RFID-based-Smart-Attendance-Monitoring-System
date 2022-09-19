import openpyxl
import time
import serial
import messagebox

pt = serial.Serial("COM4", 9600)

try:
    book_ref = openpyxl.load_workbook("reference.xlsx")
    sheet_ref = book_ref.active
except FileNotFoundError:
    book_ref = openpyxl.Workbook()
    sheet_ref = book_ref.active
    sheet_ref.cell(row=1, column=1).value = "Name"
    sheet_ref.cell(row=1, column=2).value = b"Id"
    sheet_ref.cell(row=1, column=3).value = "Time"

try:
    book_att = openpyxl.load_workbook("attended.xlsx")
    sheet_att = book_att.active
except FileNotFoundError:
    book_att = openpyxl.Workbook()
    sheet_att = book_att.active
    sheet_att.cell(row=1, column=1).value = "Name"
    sheet_att.cell(row=1, column=2).value = "Id"
    sheet_att.cell(row=1, column=3).value = "Time"


def save_students(n, fname, book, sheet):
    for i in range(n):
        name = input("enter name: ")
        id = input("enter id: ")
        now = time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime())
        sheet.append((name, id, now))
    book.save(fname)
    book.close()


def lookup_dict_ref(book, sheet, fname, key_column, value_column):
    d = {}
    for row in range(sheet.max_row):
        name = sheet.cell(row=row + 1, column=key_column).value
        uid = sheet.cell(row=row + 1, column=value_column).value
        d[uid] = name
    try:
        book.save(fname)
    except PermissionError:
        messagebox.showerror(title="Oops!!!", message="Try again after closing reference.xlsx file")
        exit(1)
    book.close()
    return d


def lookup_dict_att(book, sheet, fname, key_column, value_column):
    d = {}
    for row in range(sheet.max_row):
        name = sheet.cell(row=row + 1, column=key_column).value
        uid = sheet.cell(row=row + 1, column=value_column).value
        d[uid] = name
    try:
        book.save(fname)
    except PermissionError:
        messagebox.showerror(title="Oops!!!", message="Try again after closing attendance.xlsx file")
        exit(1)
    book.close()
    return d


def update(book, sheet):
    global lup_dict
    name = lup_dict[iput]
    now = time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime())
    sheet.append((name, iput, now))
    try:
        book.save("attendance.xlsx")
    except PermissionError:
        messagebox.showerror(title="Oops!!!", message="Try again after closing attendance.xlsx file")
        exit(1)
    book.close()


# save_students(5,"reference.xlsx",book_ref,sheet_ref)
while True:
    lup_dict = lookup_dict_ref(book_ref, sheet_ref, "reference.xlsx", 1, 2)

    iput = pt.readline().strip()
    iput = f"{iput}"

    attended_dict = lookup_dict_att(book_att, sheet_att, "attendance.xlsx", 1, 2)
    if iput in attended_dict:
        print("already scanned")
        messagebox.showinfo(title="already scanned", message="Already scanned")
    else:
        if iput in lup_dict:
            update(book_att, sheet_att)
            print("registered successfully")
            messagebox.showinfo(title="success", message="Registered successfully")
        else:
            print("not authorised")
            messagebox.showwarning(title="not authorised", message="Not authorised")
